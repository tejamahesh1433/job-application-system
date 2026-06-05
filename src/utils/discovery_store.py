"""
Job Discovery Store — persistent JSON vault for all discovered jobs.

Every job that passes through Job Discovery is stored here permanently,
even after being dismissed (X) or applied to. The status lifecycle is:

  active → applied → interview → offer → rejected
                              ↘ ghosted
  active → removed  (user hit X — still stored, just hidden from inbox)

Credit-saving dedup: before calling JSearch, we check whether we already
have fresh results for the same keyword stored in the vault. If yes, we
return them without touching the API.
"""
from __future__ import annotations

import csv
import hashlib
import io
import json
from datetime import datetime, timedelta, timezone
from pathlib import Path
from threading import RLock
from typing import Any, Dict, List, Optional

PROJECT_ROOT = Path(__file__).resolve().parents[2]
STORE_PATH   = PROJECT_ROOT / "data" / "job_discovery_inbox.json"

# All valid statuses — order matters for the UI pipeline display
ALL_STATUSES = [
    "active",
    "saved",
    "applied",
    "interview",
    "offer",
    "rejected",
    "ghosted",
    "removed",
]

STATUS_LABELS = {
    "active":    "Active",
    "saved":     "Saved",
    "applied":   "Applied",
    "interview": "Interview",
    "offer":     "Offer",
    "rejected":  "Rejected",
    "ghosted":   "Ghosted",
    "removed":   "Removed",
}

KEYWORD_FRESH_HOURS = 6


class DiscoveryStore:
    def __init__(self, path: Path = STORE_PATH):
        self.path = path
        self._lock = RLock()

    def _now(self) -> str:
        return datetime.now(timezone.utc).isoformat()

    def _load(self) -> Dict[str, Any]:
        if not self.path.exists():
            return {"jobs": {}, "searches": {}}
        try:
            data = json.loads(self.path.read_text(encoding="utf-8"))
            if not isinstance(data, dict):
                return {"jobs": {}, "searches": {}}
            data.setdefault("jobs", {})
            data.setdefault("searches", {})
            return data
        except Exception:
            return {"jobs": {}, "searches": {}}

    def _save(self, data: Dict[str, Any]) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        tmp = self.path.with_suffix(".tmp")
        tmp.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
        tmp.replace(self.path)

    def job_key(self, job: Dict[str, Any]) -> str:
        raw = (job.get("url") or "").strip().lower()
        if not raw:
            raw = f"{job.get('company','')}|{job.get('title','')}|{job.get('source','')}".lower()
        return hashlib.sha256(raw.encode()).hexdigest()[:24]

    def search_key(self, params: Dict[str, Any]) -> str:
        normalized = {
            str(k): str(v or "").strip().lower()
            for k, v in sorted(params.items())
            if k not in {"force", "_"}
        }
        raw = json.dumps(normalized, sort_keys=True, separators=(",", ":"))
        return hashlib.sha256(raw.encode()).hexdigest()[:24]

    # ── credit-saving dedup ───────────────────────────────────────────────────

    def keyword_in_vault(self, keyword: str, user_id: int = 1) -> bool:
        """
        Return True if the vault already contains ANY jobs for this keyword —
        permanently, with no time limit. This prevents re-searching keywords
        we have already paid credits for.

        The user must click 'Refresh Search' (force=True) to bypass this and
        fetch new results from JSearch for a keyword we already know.
        """
        kw = (keyword or "").strip().lower()
        if not kw:
            return False
        with self._lock:
            data = self._load()
            for record in data["jobs"].values():
                if int(record.get("user_id", 1)) != int(user_id):
                    continue
                stored_kw = (record.get("search_keyword") or "").strip().lower()
                title = ((record.get("job") or {}).get("title") or "").lower()
                # Match if keyword appears in the stored search keyword or job title
                if kw in stored_kw or kw in title:
                    return True
        return False

    # Keep old name as alias for backwards compatibility
    def keyword_has_fresh_results(self, keyword: str, user_id: int = 1) -> bool:
        return self.keyword_in_vault(keyword, user_id=user_id)

    def get_known_urls(self, user_id: int = 1) -> set:
        """
        Return the set of ALL job URLs already in the vault (any status).
        Used to filter JSearch results so we never show a job the user has
        already seen, even if they hit X or applied to it.
        """
        with self._lock:
            data = self._load()
            urls = set()
            for record in data["jobs"].values():
                if int(record.get("user_id", 1)) != int(user_id):
                    continue
                url = ((record.get("job") or {}).get("url") or "").strip().lower()
                if url:
                    urls.add(url)
            return urls

    def has_search(self, search_key: str, user_id: int = 1) -> bool:
        with self._lock:
            s = self._load()["searches"].get(search_key)
            return bool(s and int(s.get("user_id", 1)) == int(user_id))

    # ── read ──────────────────────────────────────────────────────────────────

    def _build_job_out(self, key: str, record: Dict[str, Any]) -> Dict[str, Any]:
        status = record.get("discovery_status", "active")
        job = dict(record.get("job") or {})
        job.update({
            "id":               key,
            "discovery_status": status,
            "status_label":     STATUS_LABELS.get(status, status.title()),
            "saved_at":         record.get("saved_at"),
            "last_seen_at":     record.get("last_seen_at"),
            "updated_at":       record.get("updated_at"),
            "notes":            record.get("notes", ""),
            "status_history":   record.get("status_history", []),
        })
        return job

    def list_jobs(
        self,
        user_id: int = 1,
        include_removed: bool = False,
        include_applied: bool = False,
    ) -> List[Dict[str, Any]]:
        """Active inbox view."""
        with self._lock:
            data = self._load()
            jobs = []
            for key, record in data["jobs"].items():
                if int(record.get("user_id", 1)) != int(user_id):
                    continue
                status = record.get("discovery_status", "active")
                if status == "removed" and not include_removed:
                    continue
                if status in ("applied","interview","offer","rejected","ghosted") and not include_applied:
                    continue
                jobs.append(self._build_job_out(key, record))
            jobs.sort(key=lambda j: j.get("last_seen_at") or j.get("saved_at") or "", reverse=True)
            return jobs

    def vault_list(self, user_id: int = 1, status_filter: Optional[str] = None) -> List[Dict[str, Any]]:
        """All jobs in vault, optionally filtered by status."""
        with self._lock:
            data = self._load()
            jobs = []
            for key, record in data["jobs"].items():
                if int(record.get("user_id", 1)) != int(user_id):
                    continue
                status = record.get("discovery_status", "active")
                if status_filter and status_filter != "all" and status != status_filter:
                    continue
                jobs.append(self._build_job_out(key, record))
            jobs.sort(
                key=lambda j: j.get("updated_at") or j.get("last_seen_at") or j.get("saved_at") or "",
                reverse=True,
            )
            return jobs

    def vault_counts(self, user_id: int = 1) -> Dict[str, int]:
        """Count per status + total, for badges and tab pills."""
        with self._lock:
            data = self._load()
            counts: Dict[str, int] = {s: 0 for s in ALL_STATUSES}
            counts["all"] = 0
            for record in data["jobs"].values():
                if int(record.get("user_id", 1)) != int(user_id):
                    continue
                s = record.get("discovery_status", "active")
                counts[s] = counts.get(s, 0) + 1
                counts["all"] += 1
            return counts

    def save_search(
        self,
        search_key: str,
        params: Dict[str, Any],
        jobs: List[Dict[str, Any]],
        user_id: int = 1,
        keyword: str = "",
    ) -> List[Dict[str, Any]]:
        """Persist search results. Existing status/history is never overwritten."""
        with self._lock:
            data = self._load()
            now = self._now()
            ids = []
            for job in jobs:
                key = self.job_key(job)
                existing = data["jobs"].get(key, {})
                record = {
                    "user_id":          int(user_id),
                    "discovery_status": existing.get("discovery_status", "active"),
                    "saved_at":         existing.get("saved_at", now),
                    "last_seen_at":     now,
                    "updated_at":       existing.get("updated_at", now),
                    "notes":            existing.get("notes", ""),
                    "status_history":   existing.get("status_history", []),
                    "search_keyword":   keyword or params.get("keyword", ""),
                    "job":              {**job, "id": key},
                }
                data["jobs"][key] = record
                ids.append(key)
            data["searches"][search_key] = {
                "user_id":     int(user_id),
                "params":      params,
                "job_ids":     ids,
                "searched_at": now,
            }
            self._save(data)
            return self.list_jobs(user_id=user_id)

    def update_status(
        self,
        job_id: str,
        status: str,
        user_id: int = 1,
        notes: str = "",
    ) -> Optional[Dict[str, Any]]:
        """Update status and append to history log."""
        if status not in ALL_STATUSES:
            return None
        with self._lock:
            data = self._load()
            record = data["jobs"].get(job_id)
            if not record or int(record.get("user_id", 1)) != int(user_id):
                return None
            now = self._now()
            old = record.get("discovery_status", "active")
            history = record.get("status_history", [])
            history.append({"from": old, "to": status, "at": now, "notes": notes})
            record["discovery_status"] = status
            record["updated_at"] = now
            record["status_history"] = history
            if notes:
                record["notes"] = notes
            data["jobs"][job_id] = record
            self._save(data)
            return self._build_job_out(job_id, record)

    def update_notes(self, job_id: str, notes: str, user_id: int = 1) -> bool:
        with self._lock:
            data = self._load()
            record = data["jobs"].get(job_id)
            if not record or int(record.get("user_id", 1)) != int(user_id):
                return False
            record["notes"] = notes
            record["updated_at"] = self._now()
            data["jobs"][job_id] = record
            self._save(data)
            return True

    def mark_applied_by_url(self, url: str, user_id: int = 1) -> None:
        if not url:
            return
        target = url.strip().lower()
        with self._lock:
            data = self._load()
            now = self._now()
            changed = False
            for key, record in data["jobs"].items():
                if int(record.get("user_id", 1)) != int(user_id):
                    continue
                job_url = (record.get("job") or {}).get("url", "").strip().lower()
                if job_url == target:
                    old = record.get("discovery_status", "active")
                    if old != "applied":
                        history = record.get("status_history", [])
                        history.append({"from": old, "to": "applied", "at": now, "notes": "auto"})
                        record["discovery_status"] = "applied"
                        record["updated_at"] = now
                        record["status_history"] = history
                        data["jobs"][key] = record
                        changed = True
            if changed:
                self._save(data)

    def export_csv(self, user_id: int = 1) -> str:
        jobs = self.vault_list(user_id=user_id)
        output = io.StringIO()
        fields = [
            "title", "company", "location", "status", "source",
            "salary", "work_type", "job_type", "match", "ats",
            "url", "saved_at", "updated_at", "notes",
        ]
        writer = csv.DictWriter(output, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for j in jobs:
            writer.writerow({
                "title":      j.get("title", ""),
                "company":    j.get("company", ""),
                "location":   j.get("location", ""),
                "status":     j.get("discovery_status", "active"),
                "source":     j.get("source", ""),
                "salary":     j.get("salary", ""),
                "work_type":  j.get("work_type", ""),
                "job_type":   j.get("job_type", ""),
                "match":      j.get("match", ""),
                "ats":        j.get("ats", ""),
                "url":        j.get("url", ""),
                "saved_at":   (j.get("saved_at") or "")[:10],
                "updated_at": (j.get("updated_at") or "")[:10],
                "notes":      j.get("notes", ""),
            })
        return output.getvalue()

    def export_excel_bytes(self, user_id: int = 1) -> bytes:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise RuntimeError("openpyxl not installed — run: pip install openpyxl")

        jobs = self.vault_list(user_id=user_id)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Job Vault"
        headers = [
            "Title", "Company", "Location", "Status", "Source",
            "Salary", "Work Type", "Job Type", "Match%", "ATS%",
            "Apply URL", "Discovered", "Last Updated", "Notes",
        ]
        hfill = PatternFill("solid", fgColor="1a1a2e")
        hfont = Font(bold=True, color="FFFFFF")
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = hfill
            c.font = hfont
            c.alignment = Alignment(horizontal="center")
        status_colors = {
            "active": "E8F5E9", "saved": "E3F2FD", "applied": "FFF3E0",
            "interview": "F3E5F5", "offer": "C8E6C9", "rejected": "FFEBEE",
            "ghosted": "F5F5F5", "removed": "FAFAFA",
        }
        for row, j in enumerate(jobs, 2):
            status = j.get("discovery_status", "active")
            fill = PatternFill("solid", fgColor=status_colors.get(status, "FFFFFF"))
            vals = [
                j.get("title",""), j.get("company",""), j.get("location",""),
                STATUS_LABELS.get(status, status.title()), j.get("source",""),
                j.get("salary",""), j.get("work_type",""), j.get("job_type",""),
                j.get("match",""), j.get("ats",""), j.get("url",""),
                (j.get("saved_at") or "")[:10], (j.get("updated_at") or "")[:10],
                j.get("notes",""),
            ]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.fill = fill
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(w + 4, 55)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()


discovery_store = DiscoveryStore()
