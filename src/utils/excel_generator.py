"""
Excel Generator - Create daily and master tracking spreadsheets
Color-coded by match score and performance
"""

import logging
from typing import List, Dict, Any, Optional
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelGenerator:
    """Generates tracking Excel spreadsheets"""

    def __init__(self, output_dir: str = "tracking"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate_daily_tracker(
        self,
        applications: List[Dict[str, Any]],
        date: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        Generate daily tracking Excel with 25 rows

        Columns: No., Date, Company, Job Title, Match Score, Interview %,
                 Status, Resume Version, Fields, Submitted At, Follow-up Date,
                 PDF Link, Recruiter Email, Notes
        """

        try:
            if not date:
                date = datetime.utcnow().strftime("%Y_%m_%d")

            wb = Workbook()
            ws = wb.active
            ws.title = f"Applications {date}"

            # Headers
            headers = [
                "No.", "Date", "Company", "Job Title", "Match Score",
                "Interview %", "Status", "Resume Ver", "Fields", "Submitted At",
                "Follow-up Date", "PDF Link", "Recruiter Email", "Notes"
            ]

            ws.append(headers)

            # Format header row
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Add applications
            for idx, app in enumerate(applications[:25], 1):  # Max 25
                match_score = app.get("match_score", 0)

                # Color-code by match score
                if match_score >= 85:
                    fill_color = "C6EFCE"  # Green
                elif match_score >= 70:
                    fill_color = "FFEB9C"  # Yellow
                else:
                    fill_color = "FFC7CE"  # Red

                row_data = [
                    idx,
                    app.get("date", ""),
                    app.get("company_name", ""),
                    app.get("job_title", ""),
                    f"{match_score:.0f}%",
                    app.get("interview_probability", ""),
                    app.get("status", ""),
                    app.get("resume_version", ""),
                    f"{app.get('form_fields_filled', 0)}/{app.get('form_fields_total', 0)}",
                    app.get("submitted_at", ""),
                    app.get("follow_up_date", ""),
                    app.get("pdf_path", ""),
                    app.get("recruiter_email", ""),
                    app.get("notes", ""),
                ]

                ws.append(row_data)

                # Color the row
                for cell in ws[idx + 1]:
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            # Set column widths
            column_widths = [5, 12, 15, 20, 12, 12, 12, 10, 10, 12, 12, 15, 18, 20]
            for idx, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(idx)].width = width

            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'),
            )

            for row in ws.iter_rows(min_row=1, max_row=len(applications) + 1):
                for cell in row:
                    cell.border = thin_border

            filename = f"Applications_{date}.xlsx"
            filepath = self.output_dir / filename

            wb.save(filepath)

            logger.info(f"✅ Daily tracker created: {filepath}")

            return {
                "success": True,
                "filepath": str(filepath),
                "filename": filename,
                "applications_count": len(applications[:25]),
            }

        except Exception as e:
            logger.error(f"Error generating daily tracker: {e}")
            return {"success": False, "message": str(e)}

    def generate_master_tracker(
        self,
        applications: List[Dict[str, Any]],
    ) -> Dict[str, Any]:
        """Generate cumulative master tracker with statistics"""

        try:
            wb = Workbook()

            # Sheet 1: All Applications
            ws = wb.active
            ws.title = "All Applications"

            headers = [
                "No.", "Date", "Company", "Job Title", "Match Score",
                "Status", "Response Received", "Interview Scheduled",
                "Offer Received", "Submitted At", "Response Date", "Notes"
            ]

            ws.append(headers)

            # Format header
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Add all applications
            for idx, app in enumerate(applications, 1):
                row_data = [
                    idx,
                    app.get("created_at", "")[:10],
                    app.get("company_name", ""),
                    app.get("job_title", ""),
                    f"{app.get('match_score', 0):.0f}%",
                    app.get("status", ""),
                    "Yes" if app.get("response_received") else "No",
                    "Yes" if app.get("interview_scheduled") else "No",
                    "Yes" if app.get("status") == "offer_received" else "No",
                    app.get("submitted_at", "")[:10] if app.get("submitted_at") else "",
                    app.get("response_date", "")[:10] if app.get("response_date") else "",
                    app.get("notes", ""),
                ]

                ws.append(row_data)

                # Color by status
                status = app.get("status", "")
                if status == "offer_received":
                    fill_color = "00B050"  # Green
                elif status == "interview_scheduled":
                    fill_color = "FFC000"  # Orange
                elif status == "rejected":
                    fill_color = "FF0000"  # Red
                else:
                    fill_color = "D9E1F2"  # Light blue

                for cell in ws[idx + 1]:
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

            # Set column widths
            for idx, width in enumerate([5, 12, 15, 20, 12, 12, 12, 12, 12, 12, 12, 20], 1):
                ws.column_dimensions[get_column_letter(idx)].width = width

            # Sheet 2: Statistics
            stats_ws = wb.create_sheet("Statistics")

            total = len(applications)
            submitted = sum(1 for a in applications if a.get("submitted_at"))
            responses = sum(1 for a in applications if a.get("response_received"))
            interviews = sum(1 for a in applications if a.get("interview_scheduled"))
            offers = sum(1 for a in applications if a.get("status") == "offer_received")

            stats_data = [
                ["Metric", "Count", "Percentage"],
                ["Total Applications", total, "100%"],
                ["Submitted", submitted, f"{(submitted/total*100):.1f}%" if total > 0 else "0%"],
                ["Responses", responses, f"{(responses/submitted*100):.1f}%" if submitted > 0 else "0%"],
                ["Interviews", interviews, f"{(interviews/submitted*100):.1f}%" if submitted > 0 else "0%"],
                ["Offers", offers, f"{(offers/submitted*100):.1f}%" if submitted > 0 else "0%"],
            ]

            for row in stats_data:
                stats_ws.append(row)

            # Format stats
            for cell in stats_ws[1]:
                cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")

            for idx, width in enumerate([20, 15, 15], 1):
                stats_ws.column_dimensions[get_column_letter(idx)].width = width

            filename = "Master_All_Applications.xlsx"
            filepath = self.output_dir / filename

            wb.save(filepath)

            logger.info(f"✅ Master tracker created: {filepath}")

            return {
                "success": True,
                "filepath": str(filepath),
                "filename": filename,
                "total_applications": total,
                "statistics": {
                    "submitted": submitted,
                    "responses": responses,
                    "interviews": interviews,
                    "offers": offers,
                }
            }

        except Exception as e:
            logger.error(f"Error generating master tracker: {e}")
            return {"success": False, "message": str(e)}
